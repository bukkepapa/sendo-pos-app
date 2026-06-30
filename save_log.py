# -*- coding: utf-8 -*-
import glob, os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# デスクトップ優先、次にsessions内を探す
paths = (
    glob.glob(r"C:\Users\PC_User\Desktop\Claude利用ログ.xlsx") +
    glob.glob(r"C:\Users\PC_User\AppData\Roaming\Claude\**\Claude利用ログ.xlsx", recursive=True)
)
if not paths:
    print("❌ Claude利用ログ.xlsx が見つかりません")
    raise SystemExit(1)

EXCEL_PATH = paths[0]
print(f"📄 使用ファイル: {EXCEL_PATH}")

TODAY = "2026/04/17"
chat_text = ""
cowork_text = ""
code_text = (
    "📁 作業フォルダ: C:\\Users\\PC_User\\Documents\\Claude\\Projects\\せんどうPOS分析プロジェクト\\sendo-pos-app\n"
    "・新CSVフォーマット対応（9列サマリー形式：対象年月〜点数）\n"
    "・DBにmaker_name列追加、旧不要列削除（Supabaseマイグレーション）\n"
    "・メーカー別シェア円グラフ・棒グラフページ新設（/maker）\n"
    "・CSVエンコーディング自動検出（UTF-8/Shift_JIS判定）追加\n"
    "・全ページ見出しをtext-gray-950に変更（モバイル可読性改善）\n"
    "・ダッシュボードKPI「取扱商品数」→「伊藤園商品SKU」に変更\n"
    "・前年同月比(%)表示追加（+青/-赤、箱ルール対応商品名正規化マッチング）\n"
    "・全ページで連続月合算表示対応（MonthRangeSelector）\n"
    "・AppStateContextでページ間の対象月引継ぎを実装\n"
    "・メーカーシェアページに全格納月の時系列折れ線グラフ追加\n"
    "・feature/yoy-range-timeseriesブランチ作成→GitHub push→Vercelデプロイ完了\n"
    "・GitHubリポジトリ新規作成: https://github.com/bukkepapa/sendo-pos-app"
)
memo_text = ""

wb = load_workbook(EXCEL_PATH)
ws = wb["📅 日次ログ"]

def make_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

COL_BG = {3: "D5E8D4", 4: "DAE8FC", 5: "FFE6CC", 6: "F8CECC"}

found = False
for row in ws.iter_rows(min_row=4):
    if row[0].value == TODAY:
        found = True
        texts = [chat_text, cowork_text, code_text, memo_text]
        for i, text in enumerate(texts):
            c = row[i + 2]
            if text:
                existing = c.value or ""
                c.value = (existing + "\n" + text).strip() if existing else text
                c.font = Font(name="Meiryo UI", size=10, color="1A1A1A")
                c.fill = PatternFill("solid", fgColor=COL_BG[i+3])
                c.alignment = Alignment(horizontal="left", vertical="top",
                                        wrap_text=True, indent=1)
                c.border = make_border()
        line_count = max(len([l for l in code_text.split("\n") if l.strip()]), 1)
        ws.row_dimensions[row[0].row].height = max(45, 18 * line_count)
        break

if not found:
    print(f"⚠️ {TODAY} の行がExcelに見つかりませんでした")
else:
    wb.save(EXCEL_PATH)
    print(f"✅ {TODAY} のログを保存しました")
